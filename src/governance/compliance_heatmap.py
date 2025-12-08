"""
AI4OHS-HYBRID
ESS + 6331 Compliance Heatmap Report

Amaç:
- ESS1–ESS10 ve 6331 maddeleri için uyum (E/K/H) matrisini üretmek
- Excel veya basit text tabanlı heatmap raporu üretmek
"""

from typing import Dict, List

HEATMAP_COLORS = {"E": "00B050", "K": "FFC000", "H": "FF0000"}  # Green  # Orange  # Red


def generate_heatmap_matrix(items: List[Dict]) -> Dict[str, Dict[str, str]]:
    """
    items: [
      {"ess": "ESS2", "law_ref": "6331-4", "status": "E"},
      ...
    ]

    Dönen:
      { "ESS2": { "6331-4": "E", "6331-10": "H", ... }, ... }
    """
    matrix = {}
    for it in items:
        ess = it["ess"]
        law = it["law_ref"]
        status = (it.get("status") or "H").upper()

        matrix.setdefault(ess, {})
        matrix[ess][law] = status

    return matrix


def export_heatmap_excel(matrix: Dict[str, Dict[str, str]], filename="ESS_6331_Heatmap.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl dependency missing for heatmap export.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "ESS_6331_Heatmap"

    # kolon başlıkları: ESS sütun, 6331 satır
    all_laws = sorted({law for vals in matrix.values() for law in vals.keys()})
    ess_list = sorted(matrix.keys())

    ws.cell(row=1, column=1, value="6331 / ESS")
    for j, ess in enumerate(ess_list, start=2):
        ws.cell(row=1, column=j, value=ess)

    for i, law in enumerate(all_laws, start=2):
        ws.cell(row=i, column=1, value=law)
        for j, ess in enumerate(ess_list, start=2):
            status = matrix.get(ess, {}).get(law, "")
            cell = ws.cell(row=i, column=j, value=status)
            if status in HEATMAP_COLORS:
                cell.fill = PatternFill("solid", fgColor=HEATMAP_COLORS[status])

    wb.save(filename)
    return filename


def render_heatmap_text(matrix: Dict[str, Dict[str, str]]) -> str:
    """
    Basit ASCII heatmap özet yazımı (CLI raporları için).
    """
    ess_list = sorted(matrix.keys())
    all_laws = sorted({law for vals in matrix.values() for law in vals.keys()})

    lines = []
    header = "Law/ESS | " + " | ".join(ess_list)
    lines.append(header)
    lines.append("-" * len(header))

    for law in all_laws:
        row = [law]
        for ess in ess_list:
            status = matrix.get(ess, {}).get(law, "")
            row.append(status or "-")
        lines.append(" | ".join(row))

    return "\n".join(lines)
